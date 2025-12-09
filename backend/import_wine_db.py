import openpyxl
import asyncio
import sys
import os
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timezone
import uuid
from emergentintegrations.llm.chat import LlmChat, UserMessage

# Add parent directory to path
sys.path.insert(0, os.path.dirname(__file__))

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'wine_pairing_db')]

# LLM API Key
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY', '')


def extract_wine_data_from_excel(file_path):
    """Extract all wine data from all sheets in the Excel file."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_wines = []
    
    print(f"📊 Found {len(wb.sheetnames)} sheets in Excel file")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n🔍 Processing sheet: {sheet_name}")
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        if not headers:
            print(f"  ⚠️  Skipping empty sheet")
            continue
        
        print(f"  📋 Headers: {headers}")
        
        # Process rows
        row_count = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Create a dictionary from row data
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and value:
                    row_data[headers[idx]] = str(value).strip()
            
            if not row_data:
                continue
            
            # Extract wine information based on available fields
            wine_entry = extract_wine_from_row(row_data, sheet_name)
            if wine_entry:
                all_wines.append(wine_entry)
                row_count += 1
        
        print(f"  ✅ Extracted {row_count} wines")
    
    print(f"\n🎉 Total wines extracted: {len(all_wines)}")
    return all_wines


def extract_wine_from_row(row_data, sheet_name):
    """Extract wine information from a single row."""
    
    # Determine wine name - try multiple column names
    name = None
    for key in ['Château', 'Château / Domaine', 'Premier Cru', 'Cru', 'Wein', 'Appellation', 'MGA-Lage (Gemeinde)', 'Grand Cru (Lage)']:
        if key in row_data:
            name = row_data[key]
            break
    
    if not name:
        return None
    
    # Determine region
    region = row_data.get('Region', row_data.get('Appellation', 'Bordeaux'))
    
    # Determine description
    description = row_data.get('Beschreibung', row_data.get('Emotionaler Charakter', row_data.get('Stil-Beschreibung', '')))
    
    if not description:
        return None
    
    # Determine grape variety
    grape_variety = row_data.get('Rebsorten-DNA (Dominant)', row_data.get('Rebsorten-DNA', row_data.get('Seele des Weins (Rebsorten)', row_data.get('Dominante Rebsorten', 'Unbekannt'))))
    
    # Determine food pairings
    pairings = row_data.get('Das perfekte Pairing', '')
    pairings_list = [p.strip() for p in pairings.split(',') if p.strip()] if pairings else []
    
    # Determine wine color from sheet name or region
    wine_color = determine_wine_color(sheet_name, region, name)
    
    # Determine country from region/sheet context
    country = determine_country(sheet_name, region)
    
    wine_entry = {
        'id': str(uuid.uuid4()),
        'name': name,
        'winery': name,  # In most cases, château name = wine name
        'country': country,
        'region': region,
        'appellation': region,
        'grape_variety': grape_variety,
        'wine_color': wine_color,
        'year': None,
        'description_de': description,
        'description_en': None,  # To be generated
        'description_fr': None,  # To be generated
        'tasting_notes': row_data.get('Stil-Beschreibung', None),
        'food_pairings_de': pairings_list,
        'food_pairings_en': [],  # To be generated
        'food_pairings_fr': [],  # To be generated
        'alcohol_content': None,
        'price_category': determine_price_category(sheet_name, name),
        'image_url': '/placeholder-wine.png',
        'rating': None,
        'created_at': datetime.now(timezone.utc)
    }
    
    return wine_entry


def determine_wine_color(sheet_name, region, name):
    """Determine wine color based on context."""
    sheet_lower = sheet_name.lower()
    name_lower = name.lower()
    region_lower = region.lower()
    
    # Check for white wine indicators
    if any(word in sheet_lower or word in name_lower or word in region_lower for word in 
           ['weiß', 'white', 'blanc', 'chardonnay', 'riesling', 'sauvignon blanc', 'chenin', 'viognier', 'grüner veltliner']):
        return 'weiss'
    
    # Check for sweet wine indicators
    if any(word in sheet_lower or word in name_lower or word in region_lower for word in 
           ['sauternes', 'barsac', 'süß', 'sweet', 'vendanges tardives', 'sélection de grains nobles']):
        return 'suesswein'
    
    # Check for sparkling wine indicators
    if any(word in sheet_lower or word in name_lower or word in region_lower for word in 
           ['champagner', 'champagne', 'sekt', 'crémant', 'prosecco', 'cava']):
        return 'schaumwein'
    
    # Check for rosé indicators
    if any(word in sheet_lower or word in name_lower or word in region_lower for word in 
           ['rosé', 'rose', 'rosato', 'provence']):
        return 'rose'
    
    # Default to red wine
    return 'rot'


def determine_country(sheet_name, region):
    """Determine country based on sheet name and region."""
    context = (sheet_name + ' ' + region).lower()
    
    if any(word in context for word in ['bordeaux', 'bourgogne', 'burgund', 'rhône', 'rhone', 'loire', 'champagne', 'provence', 'alsace', 'sauternes']):
        return 'Frankreich'
    
    if any(word in context for word in ['toskana', 'tuscany', 'piemont', 'venetien', 'barolo', 'barbaresco', 'brunello', 'chianti', 'amarone']):
        return 'Italien'
    
    if any(word in context for word in ['deutschland', 'germany', 'mosel', 'rheingau', 'pfalz', 'riesling']):
        return 'Deutschland'
    
    if any(word in context for word in ['spanien', 'spain', 'rioja', 'ribera', 'priorat', 'tempranillo']):
        return 'Spanien'
    
    if any(word in context for word in ['portugal', 'douro', 'porto', 'port']):
        return 'Portugal'
    
    if any(word in context for word in ['usa', 'napa', 'california', 'sonoma']):
        return 'USA'
    
    # Default to France (most wines in the file are French)
    return 'Frankreich'


def determine_price_category(sheet_name, name):
    """Determine price category based on classification."""
    sheet_lower = sheet_name.lower()
    name_lower = name.lower()
    
    # Luxury wines
    if any(word in sheet_lower or word in name_lower for word in 
           ['premier cru classé', '1. gcc', 'grand cru', 'pétrus', 'lafite', 'latour', 'margaux', 'dom pérignon', 'cristal', 'krug']):
        return 'luxury'
    
    # Premium wines
    if any(word in sheet_lower or word in name_lower for word in 
           ['deuxièmes', '2.', 'troisièmes', '3.', 'quatrièmes', '4.', 'cinquièmes', '5.', 'barbaresco', 'barolo', 'brunello']):
        return 'premium'
    
    # Mid-range wines
    if any(word in sheet_lower for word in ['cru bourgeois', 'premier cru', 'appellation']):
        return 'mid-range'
    
    # Default to mid-range
    return 'mid-range'


async def generate_translations(wine_entry):
    """Generate English and French translations for description and pairings."""
    
    try:
        # Generate English translation
        chat_en = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=str(uuid.uuid4()),
            system_message="You are an expert wine translator. Translate wine descriptions from German to English, preserving the emotional and poetic tone."
        )
        
        prompt_en = f"""Translate the following emotional wine description to English. Keep the poetic, emotional tone:

Description (DE): {wine_entry['description_de']}

Food Pairings (DE): {', '.join(wine_entry['food_pairings_de']) if wine_entry['food_pairings_de'] else 'None'}

Return the result in this JSON format:
{{
  "description_en": "...",
  "food_pairings_en": ["...", "..."]
}}"""
        
        response_en = await chat_en.generate_response([UserMessage(content=prompt_en)])
        
        # Extract JSON from response
        import json
        import re
        json_match = re.search(r'\{.*\}', response_en, re.DOTALL)
        if json_match:
            data_en = json.loads(json_match.group())
            wine_entry['description_en'] = data_en.get('description_en', '')
            wine_entry['food_pairings_en'] = data_en.get('food_pairings_en', [])
        
        # Generate French translation
        prompt_fr = f"""Traduis cette description émotionnelle de vin en français. Garde le ton poétique et émotionnel:

Description (DE): {wine_entry['description_de']}

Food Pairings (DE): {', '.join(wine_entry['food_pairings_de']) if wine_entry['food_pairings_de'] else 'None'}

Retourne le résultat au format JSON suivant:
{{
  "description_fr": "...",
  "food_pairings_fr": ["...", "..."]
}}"""
        
        response_fr = await chat.generate_response([UserMessage(content=prompt_fr)])
        
        # Extract JSON from response
        json_match = re.search(r'\{.*\}', response_fr, re.DOTALL)
        if json_match:
            data_fr = json.loads(json_match.group())
            wine_entry['description_fr'] = data_fr.get('description_fr', '')
            wine_entry['food_pairings_fr'] = data_fr.get('food_pairings_fr', [])
        
        print(f"  ✅ Translations generated for {wine_entry['name']}")
        return True
        
    except Exception as e:
        print(f"  ⚠️  Translation failed for {wine_entry['name']}: {str(e)}")
        # Set empty translations as fallback
        wine_entry['description_en'] = wine_entry['description_de']
        wine_entry['description_fr'] = wine_entry['description_de']
        wine_entry['food_pairings_en'] = wine_entry['food_pairings_de']
        wine_entry['food_pairings_fr'] = wine_entry['food_pairings_de']
        return False


async def import_to_database(wines):
    """Import all wines to MongoDB, replacing existing data."""
    
    # Delete all existing wines in wines_db collection
    deleted_count = await db.wines_db.delete_many({})
    print(f"\n🗑️  Deleted {deleted_count.deleted_count} existing wines from wines_db")
    
    # Generate translations and insert wines
    print(f"\n🌍 Generating translations for {len(wines)} wines...")
    
    for idx, wine in enumerate(wines, 1):
        print(f"\n[{idx}/{len(wines)}] Processing: {wine['name']}")
        
        # Generate translations
        await generate_translations(wine)
        
        # Convert datetime to ISO string for MongoDB
        wine_copy = wine.copy()
        wine_copy['created_at'] = wine_copy['created_at'].isoformat()
        
        # Insert into database
        await db.wines_db.insert_one(wine_copy)
        print(f"  💾 Inserted into database")
    
    # Count final wines
    final_count = await db.wines_db.count_documents({})
    print(f"\n🎉 Import complete! Total wines in database: {final_count}")


async def main():
    print("🍷 Starting Wine Database Import\n")
    print("=" * 60)
    
    # Extract wines from Excel
    wines = extract_wine_data_from_excel('/app/wine_db_source.xlsx')
    
    if not wines:
        print("❌ No wines extracted from Excel file!")
        return
    
    # Import to database
    await import_to_database(wines)
    
    print("\n" + "=" * 60)
    print("✅ Wine Database Import Complete!")


if __name__ == '__main__':
    asyncio.run(main())
