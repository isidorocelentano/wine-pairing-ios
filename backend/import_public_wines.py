"""
Import wines from Excel into the public_wines collection

Handles TWO different column structures:
6-column format (most wines):
  [0] Wine Name, [1] Appellation/Status, [2] Classification, [3] Description, [4] Grape, [5] Pairing

7-column format (some German/Austrian wines):
  [0] Wine Name, [1] Appellation/Status, [2] Classification, [3] Keyword, [4] Description, [5] Grape, [6] Pairing

Also handles TWO appellation formats:
  Format A: "Region / Appellation_or_Class / Country" (e.g., "Südtirol / Alto Adige / Italien")
  Format B: "Country / Classification / Region" (e.g., "Deutschland / Kult-Wein / Mosel")
"""
import openpyxl
import asyncio
import os
import re
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timezone
import uuid

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'wine_pairing_db')]

# Define known countries (lowercase -> display name)
KNOWN_COUNTRIES = {
    'italien': 'Italien',
    'frankreich': 'Frankreich', 
    'spanien': 'Spanien',
    'deutschland': 'Deutschland',
    'österreich': 'Österreich',
    'schweiz': 'Schweiz',
    'portugal': 'Portugal',
    'usa': 'USA',
    'australien': 'Australien',
    'chile': 'Chile',
    'argentinien': 'Argentinien',
    'ungarn': 'Ungarn',
    'neuseeland': 'Neuseeland',
    'südafrika': 'Südafrika',
}

# Classification terms that are NOT geographic
CLASSIFICATION_TERMS = {
    # Quality classifications
    'cru', 'doc', 'docg', 'igt', 'vdit', 'do', 'doca',
    'reserva', 'crianza', 'gran reserva', 'grand cru', 'premier cru',
    'gran selezione', 'riserva', 'superiore', 'prädikatswein',
    'premier cru supérieur', 'premier cru classé', 'deuxième cru',
    'cinquième cru', 'cru classé', 'troisième cru', 'quatrième cru',
    'grosses gewächs', 'trockenbeerenauslese',
    # Wine types (not geographic)
    'rotwein', 'weißwein', 'schaumwein', 'süßwein', 'rosado', 'rosé',
    'cava', 'champagner', 'sekt', 'prosecco',
    # Status terms
    'status', 'prestige-wein', 'kult-wein', 'ikone', 'ikonen-status', 
    'basiswein', 'zweitwein', 'super tuscan',
    # Generic
    'appellation',
}


def is_country(value: str) -> bool:
    """Check if value is a known country"""
    return value.lower().strip() in KNOWN_COUNTRIES


def is_classification(value: str) -> bool:
    """Check if value is a classification term"""
    return value.lower().strip() in CLASSIFICATION_TERMS


def parse_appellation_status(value: str) -> dict:
    """Parse the 'Appellation / Status' field handling both formats."""
    if not value:
        return {'country': None, 'region': None, 'appellation': None, 'classification': None}
    
    parts = [p.strip() for p in value.split(' / ') if p.strip()]
    
    country = None
    region = None
    appellation = None
    classification = None
    
    if len(parts) == 3:
        first, second, third = parts
        
        first_is_country = is_country(first)
        third_is_country = is_country(third)
        second_is_class = is_classification(second)
        
        if first_is_country and not third_is_country:
            # Format B: Country / Classification / Region
            country = KNOWN_COUNTRIES.get(first.lower(), first)
            classification = second if second_is_class else None
            region = third
            appellation = third
            
        elif third_is_country and not first_is_country:
            # Format A: Region / Appellation_or_Class / Country
            country = KNOWN_COUNTRIES.get(third.lower(), third)
            region = first
            
            if second_is_class:
                classification = second
                appellation = first
            else:
                appellation = second
                
        else:
            # Fallback
            for i, part in enumerate(parts):
                if is_country(part):
                    country = KNOWN_COUNTRIES.get(part.lower(), part)
                    remaining = [p for j, p in enumerate(parts) if j != i]
                    if remaining:
                        region = remaining[0]
                        appellation = remaining[1] if len(remaining) > 1 else remaining[0]
                    break
            
            if not country:
                region = first
                appellation = second if not second_is_class else first
                classification = second if second_is_class else None
                
    elif len(parts) == 2:
        first, second = parts
        
        if is_country(first):
            country = KNOWN_COUNTRIES.get(first.lower(), first)
            region = second
            appellation = second
        elif is_country(second):
            country = KNOWN_COUNTRIES.get(second.lower(), second)
            region = first
            appellation = first
        else:
            region = first
            if is_classification(second):
                classification = second
                appellation = first
            else:
                appellation = second
                
    elif len(parts) == 1:
        region = parts[0]
        appellation = parts[0]
        if is_country(parts[0]):
            country = KNOWN_COUNTRIES.get(parts[0].lower(), parts[0])
            region = None
            appellation = None
    
    return {
        'country': country,
        'region': region,
        'appellation': appellation,
        'classification': classification
    }


def extract_multilingual_description(desc_text: str) -> tuple:
    """Extract German, English, and French descriptions."""
    if not desc_text:
        return ('', '', '')
    
    desc_de = desc_text.strip()
    desc_en = desc_text.strip()
    desc_fr = desc_text.strip()
    
    paren_matches = re.findall(r'\(([^)]+)\)', desc_text)
    
    if len(paren_matches) >= 2:
        desc_en = paren_matches[0]
        desc_fr = paren_matches[1]
        first_paren = desc_text.find('(')
        if first_paren > 0:
            desc_de = desc_text[:first_paren].strip()
    elif len(paren_matches) == 1:
        desc_en = paren_matches[0]
        desc_fr = paren_matches[0]
        first_paren = desc_text.find('(')
        if first_paren > 0:
            desc_de = desc_text[:first_paren].strip()
    
    return (desc_de, desc_en, desc_fr)


def determine_wine_color(wine_name: str, grape: str, region: str, classification: str) -> str:
    """Determine wine color based on various indicators"""
    context = f"{wine_name} {grape} {region} {classification or ''}".lower()
    
    white_indicators = [
        'weiss', 'weiß', 'white', 'blanc', 'bianco',
        'chardonnay', 'riesling', 'sauvignon blanc', 'pinot grigio', 'pinot gris',
        'gewürztraminer', 'moscato', 'grüner veltliner', 'albariño', 'verdejo',
        'viognier', 'chenin', 'semillon', 'trebbiano', 'vermentino', 'weißwein',
        'gruner veltliner', 'silvaner', 'müller-thurgau', 'müller thurgau'
    ]
    
    rose_indicators = ['rosé', 'rose', 'rosado', 'rosato']
    sweet_indicators = ['sauternes', 'süß', 'sweet', 'tokaji', 'eiswein', 'auslese', 
                       'beerenauslese', 'trockenbeerenauslese', 'süßwein', 'dessert',
                       'prädikatswein']
    sparkling_indicators = ['champagne', 'champagner', 'sekt', 'cava', 'prosecco', 
                           'crémant', 'spumante', 'schaumwein', 'franciacorta', 'brut']
    
    if any(ind in context for ind in sparkling_indicators):
        return 'schaumwein'
    if any(ind in context for ind in sweet_indicators):
        return 'suesswein'
    if any(ind in context for ind in rose_indicators):
        return 'rose'
    if any(ind in context for ind in white_indicators):
        return 'weiss'
    
    return 'rot'


def determine_price_category(classification: str, wine_name: str) -> str:
    """Determine price category based on classification and name"""
    context = f"{classification or ''} {wine_name}".lower()
    
    luxury_terms = ['premier cru classé', 'grand cru', 'gran selezione', 'ikone', 
                   'ikonen-status', 'kult-wein', 'pétrus', 'lafite', 'latour', 
                   'margaux', 'mouton', 'prestige-wein', 'trockenbeerenauslese']
    premium_terms = ['reserva', 'gran reserva', 'barbaresco', 'barolo', 'brunello', 
                    'super tuscan', 'crianza', 'riserva', 'cru classé', 'grosses gewächs']
    
    if any(term in context for term in luxury_terms):
        return 'luxury'
    if any(term in context for term in premium_terms):
        return 'premium'
    
    return 'mid-range'


def detect_column_format(data: list) -> str:
    """
    Detect whether this row uses 6-column or 7-column format.
    
    6-column: [name, location, class, description(long), grape(short), pairing]
    7-column: [name, location, class, keyword(short), description(long), grape(short), pairing]
    """
    if len(data) < 6:
        return 'unknown'
    
    # Check column 3 length - if it's short (< 50 chars), likely 7-column format with keyword
    col3 = data[3].strip() if len(data) > 3 else ''
    col4 = data[4].strip() if len(data) > 4 else ''
    
    # If column 3 is very short and column 4 is long, it's 7-column format
    if len(col3) < 50 and len(col4) > 100:
        return '7-column'
    
    # Standard 6-column format
    return '6-column'


def extract_wines_from_excel(file_path):
    """Extract wines from Excel with adaptive column handling."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_wines = []
    
    print(f"📊 Processing {len(wb.sheetnames)} sheets")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        header_row = sheet[1]
        header_value = header_row[0].value if header_row[0].value else ""
        
        if ';' not in str(header_value):
            print(f"\n   ⚠️ Sheet '{sheet_name}' - Not semicolon format, skipping")
            continue
            
        print(f"\n   📄 Sheet '{sheet_name}'")
        
        format_counts = {'6-column': 0, '7-column': 0}
        sheet_wines = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            
            data = str(row[0]).split(';')
            
            if len(data) < 5:
                continue
            
            wine_name = data[0].strip() if len(data) > 0 else None
            appellation_status = data[1].strip() if len(data) > 1 else None
            classification_col = data[2].strip() if len(data) > 2 else None
            
            if not wine_name:
                continue
            
            # Detect column format and extract accordingly
            col_format = detect_column_format(data)
            format_counts[col_format] = format_counts.get(col_format, 0) + 1
            
            if col_format == '7-column':
                # 7-column format: keyword in col3, description in col4
                keyword = data[3].strip() if len(data) > 3 else ''
                description = data[4].strip() if len(data) > 4 else None
                grape = data[5].strip() if len(data) > 5 else 'Unbekannt'
                pairing = data[6].strip() if len(data) > 6 else ''
                
                # Prepend keyword to description if meaningful
                if keyword and description:
                    description = f"{keyword}. {description}"
            else:
                # 6-column format: description in col3
                description = data[3].strip() if len(data) > 3 else None
                grape = data[4].strip() if len(data) > 4 else 'Unbekannt'
                pairing = data[5].strip() if len(data) > 5 else ''
            
            if not description:
                continue
            
            # Parse hierarchy
            hierarchy = parse_appellation_status(appellation_status)
            
            # Use classification from parsing or column
            classification = hierarchy.get('classification') or classification_col
            
            # Extract multilingual descriptions
            desc_de, desc_en, desc_fr = extract_multilingual_description(description)
            
            # Parse pairings (handle comma or semicolon separated)
            if pairing:
                pairings = [p.strip() for p in re.split(r'[,;]', pairing) if p.strip()]
            else:
                pairings = []
            
            # Extract winery
            winery = wine_name.split()[0] if wine_name else 'Unbekannt'
            
            # Determine color and price
            wine_color = determine_wine_color(wine_name, grape, hierarchy.get('region', ''), classification)
            price_category = determine_price_category(classification, wine_name)
            
            wine = {
                'id': str(uuid.uuid4()),
                'name': wine_name,
                'winery': winery,
                'country': hierarchy.get('country') or 'Unbekannt',
                'region': hierarchy.get('region') or 'Unbekannt',
                'appellation': hierarchy.get('appellation') or hierarchy.get('region') or 'Unbekannt',
                'grape_variety': grape if grape else 'Unbekannt',
                'wine_color': wine_color,
                'year': None,
                'description_de': desc_de,
                'description_en': desc_en,
                'description_fr': desc_fr,
                'tasting_notes': classification,
                'food_pairings_de': pairings,
                'food_pairings_en': pairings,
                'food_pairings_fr': pairings,
                'alcohol_content': None,
                'price_category': price_category,
                'image_url': '/placeholder-wine.png',
                'rating': None,
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
            all_wines.append(wine)
            sheet_wines += 1
        
        print(f"      → Extracted {sheet_wines} wines")
        print(f"      → Format breakdown: {format_counts}")
    
    print(f"\n✅ Total wines extracted: {len(all_wines)}")
    return all_wines


async def import_to_db(wines):
    """Import wines to database"""
    
    await db.public_wines.delete_many({})
    print("\n🗑️  Cleared public_wines collection")
    
    if wines:
        await db.public_wines.insert_many(wines)
        print(f"💾 Inserted {len(wines)} wines")
    
    count = await db.public_wines.count_documents({})
    print(f"✅ Final count: {count}")
    
    # Statistics
    print("\n📊 Statistics:")
    
    countries = await db.public_wines.distinct("country")
    valid_countries = sorted([c for c in countries if c != 'Unbekannt'])
    print(f"   Countries ({len(valid_countries)}): {', '.join(valid_countries)}")
    
    # Count by country
    print("\n   Wines per country:")
    for country in valid_countries:
        cnt = await db.public_wines.count_documents({"country": country})
        print(f"      {country}: {cnt}")
    
    colors = await db.public_wines.distinct("wine_color")
    print(f"\n   Wine colors: {', '.join(sorted(colors))}")
    
    # Sample hierarchy
    print("\n📍 Sample Hierarchy (various countries):")
    for country in ['Italien', 'Deutschland', 'Frankreich', 'Spanien'][:4]:
        sample = await db.public_wines.find_one(
            {"country": country}, 
            {"_id": 0, "name": 1, "country": 1, "region": 1, "appellation": 1, "grape_variety": 1, "description_de": 1}
        )
        if sample:
            name = sample['name'][:30] + '...' if len(sample['name']) > 30 else sample['name']
            desc = sample['description_de'][:50] + '...' if len(sample['description_de']) > 50 else sample['description_de']
            print(f"   [{sample['country']}] {name}")
            print(f"      Region: {sample['region']} → Appellation: {sample['appellation']}")
            print(f"      Grape: {sample['grape_variety']}")
            print(f"      Desc: {desc}")


async def main():
    print("🍷 Wine Database Import - Adaptive Column Format Handler\n")
    print("=" * 60)
    
    file_path = '/app/wine_db_gross.xlsx'
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    print(f"📂 Source: {file_path}\n")
    
    wines = extract_wines_from_excel(file_path)
    
    if wines:
        await import_to_db(wines)
    else:
        print("❌ No wines extracted!")
    
    print("\n" + "=" * 60)
    print("✅ Import Complete!")


if __name__ == '__main__':
    asyncio.run(main())
